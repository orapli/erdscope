"""Integration tests against real MySQL / PostgreSQL servers (see
tests/fixture_db/demo_mysql.sql and demo_postgres.sql).

Soft-dependency, environment-gated (same pattern as test_e2e.py's
HAVE_PLAYWRIGHT / test_erd.py's openpyxl roundtrip tests): everything here
is skipped unless ERDSCOPE_IT_MYSQL_URL / ERDSCOPE_IT_POSTGRES_URL are set,
so this file is safe to pick up in a plain `unittest discover` run with no
database available. .github/workflows/ci.yml's `db-integration` job sets
both against MySQL 8.4 / PostgreSQL 16 service containers.

Run locally against the erd-mysql / erd-pg containers described in the repo
docs:
    export ERDSCOPE_IT_MYSQL_URL="mysql://root:root@127.0.0.1:3306/erdscope_it"
    export ERDSCOPE_IT_POSTGRES_URL="postgres://postgres:pg@127.0.0.1:5432/erdscope_it"
    python3 -m unittest tests.test_db_integration -v

Each setUpClass drops and recreates the `erdscope_it` database (via the
mysql/psql CLI clients) and re-applies its fixture file, so reruns are
idempotent and don't collide with any other database on the same server.

Every real assertion drives erd.py as a subprocess exactly like a user would
(mysql://... / postgres://... URL, --excel, --no-open) and inspects the
`const DATA = ...` JSON embedded in the generated HTML — same pattern as
test_demo.py. Each schema is fetched TWICE per test: once with the PyMySQL/
psycopg driver installed, and once with a stub PYTHONPATH entry that makes
those imports raise ImportError, forcing erd.py down its dependency-free
mysql/psql CLI fallback. The two DATA JSONs must be byte-identical; this is
also the regression test for the NULL-default bug fixed in
src/erdscope/db/base.py (_unescape_mysql_field): the mysql `--batch` client
spells SQL NULL as the bare word `NULL` (`\\N` is the INTO OUTFILE / mysqldump
spelling, not the batch client's), which used to leak into the IR as the
literal string 'NULL' on the CLI path only.
"""
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from urllib.parse import urlparse

ROOT = Path(__file__).resolve().parent.parent
FIXTURE_DIR = Path(__file__).resolve().parent / 'fixture_db'

MYSQL_URL = os.environ.get('ERDSCOPE_IT_MYSQL_URL')
POSTGRES_URL = os.environ.get('ERDSCOPE_IT_POSTGRES_URL')

_STUB_SOURCE = 'raise ImportError("forced for CLI-fallback integration test")\n'
_STUB_DIR = None  # set in setUpModule, torn down in tearDownModule


def setUpModule():
    global _STUB_DIR
    if not (MYSQL_URL or POSTGRES_URL):
        return  # nothing gated on this module will run; skip the setup too
    _STUB_DIR = tempfile.mkdtemp(prefix='erdscope_it_stub_')
    for name in ('pymysql', 'psycopg', 'psycopg2'):
        (Path(_STUB_DIR) / f'{name}.py').write_text(_STUB_SOURCE, encoding='utf-8')


def tearDownModule():
    if _STUB_DIR:
        shutil.rmtree(_STUB_DIR, ignore_errors=True)


def _run_erd(url, out, extra_args=(), cli_fallback=False):
    """Drive erd.py as a subprocess exactly as a user's shell would.
    cli_fallback=True inserts the stub PYTHONPATH entry (see setUpModule)
    so pymysql/psycopg/psycopg2 all fail to import and the DB adapters fall
    back to the mysql/psql CLI clients."""
    env = dict(os.environ)
    if cli_fallback:
        assert _STUB_DIR, 'setUpModule did not run'
        env['PYTHONPATH'] = _STUB_DIR + os.pathsep + env.get('PYTHONPATH', '')
    cmd = [sys.executable, str(ROOT / 'erd.py'), url, '--no-open', '-o', str(out), *extra_args]
    return subprocess.run(cmd, capture_output=True, text=True, env=env)


def _load_data(html_path):
    html = Path(html_path).read_text(encoding='utf-8')
    m = re.search(r'^const DATA = (.+);$', html, re.MULTILINE)
    assert m, 'const DATA = ...; not found in generated HTML'
    return json.loads(m.group(1))


def _assert_no_stray_null_strings(tables):
    """Regression guard for the mysql-CLI NULL-unescaping bug: no column
    should ever carry the literal string 'NULL' as its default — a real
    absent default is either no `default` key at all or an empty one,
    never the four-letter word."""
    for tname, t in tables.items():
        for c in t['columns']:
            assert c.get('default') != 'NULL', (
                f'{tname}.{c["name"]} has default \'NULL\' as a literal string '
                '— the mysql CLI NULL-unescaping bug has regressed')


@unittest.skipUnless(MYSQL_URL, 'set ERDSCOPE_IT_MYSQL_URL to run MySQL integration tests, '
                                'e.g. mysql://root:root@127.0.0.1:3306/erdscope_it')
class TestMySQLIntegration(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        u = urlparse(MYSQL_URL)
        db = u.path.lstrip('/')
        if not re.fullmatch(r'\w+', db or ''):
            raise RuntimeError(f'ERDSCOPE_IT_MYSQL_URL must include a plain database name, got {db!r}')
        if not re.fullmatch(r'\w+', db or '') or not db.startswith('erdscope'):
            raise RuntimeError(
                'ERDSCOPE_IT_MYSQL_URL must name a disposable database whose name '
                f'starts with "erdscope" — it is DROPPED and recreated. Got {db!r}')
        cls.db = db
        host, port, user = u.hostname or '127.0.0.1', u.port or 3306, u.username or 'root'
        env = dict(os.environ)
        if u.password:
            env['MYSQL_PWD'] = u.password
        base = ['mysql', '-h', host, '-P', str(port), '-u', user]
        subprocess.run(base + ['-e', f'DROP DATABASE IF EXISTS {db}; CREATE DATABASE {db};'],
                        check=True, env=env, capture_output=True, text=True)
        fixture_sql = (FIXTURE_DIR / 'demo_mysql.sql').read_text(encoding='utf-8')
        subprocess.run(base + [db], input=fixture_sql, check=True, env=env,
                        capture_output=True, text=True)

    def test_schema_matches_across_driver_and_cli_paths(self):
        with tempfile.TemporaryDirectory() as tmp:
            out_driver = Path(tmp) / 'driver.html'
            out_cli = Path(tmp) / 'cli.html'
            r1 = _run_erd(MYSQL_URL, out_driver)
            self.assertEqual(r1.returncode, 0, r1.stderr)
            r2 = _run_erd(MYSQL_URL, out_cli, cli_fallback=True)
            self.assertEqual(r2.returncode, 0, r2.stderr)
            data_driver = _load_data(out_driver)
            data_cli = _load_data(out_cli)
            _assert_no_stray_null_strings(data_cli['tables'])
            self.assertEqual(data_driver, data_cli,
                'PyMySQL driver path and mysql-CLI fallback path must produce '
                'byte-identical IR (see _unescape_mysql_field NULL handling)')
            self._assert_schema(data_driver['tables'])

    def _assert_schema(self, tables):
        self.assertEqual(len(tables), 13)
        self.assertNotIn('active_users', tables)  # VIEW must be excluded

        users = tables['users']
        self.assertEqual(users['comment'], 'ユーザー')
        name_col = next(c for c in users['columns'] if c['name'] == 'name')
        self.assertEqual(name_col['comment'], '氏名')

        addresses = tables['addresses']
        self.assertEqual(addresses['comment'], '住所\tタブ入りコメント')

        # shipments.order_id is UNIQUE, so the DB FK promotes to 1:1
        ship_assocs = tables['shipments']['associations']
        self.assertEqual(len(ship_assocs), 1)
        self.assertEqual(ship_assocs[0]['type'], 'has_one')
        self.assertEqual(ship_assocs[0]['target'], 'orders')
        self.assertEqual(ship_assocs[0]['foreign_key'], 'order_id')
        self.assertTrue(ship_assocs[0]['db_fk'])

        # composite primary key
        pc_cols = {c['name']: c for c in tables['product_categories']['columns']}
        self.assertTrue(pc_cols['product_id']['primary'])
        self.assertTrue(pc_cols['category_id']['primary'])

        # composite index
        addr_idx = {ix['name']: ix for ix in addresses['indexes']}
        self.assertEqual(addr_idx['idx_addresses_user_kind']['columns'], ['user_id', 'kind'])

    def test_excel_export_generated(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'out.html'
            xlsx = Path(tmp) / 'out.xlsx'
            r = _run_erd(MYSQL_URL, out, extra_args=['--excel', str(xlsx)])
            self.assertEqual(r.returncode, 0, r.stderr)
            self.assertTrue(out.exists())
            self.assertTrue(xlsx.exists())
            import openpyxl  # optional dep; installed alongside pymysql in CI
            wb = openpyxl.load_workbook(xlsx)
            self.assertIn('users', wb.sheetnames)


@unittest.skipUnless(POSTGRES_URL, 'set ERDSCOPE_IT_POSTGRES_URL to run PostgreSQL integration '
                                   'tests, e.g. postgres://postgres:pg@127.0.0.1:5432/erdscope_it')
class TestPostgresIntegration(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        u = urlparse(POSTGRES_URL)
        db = u.path.lstrip('/')
        if not re.fullmatch(r'\w+', db or ''):
            raise RuntimeError(f'ERDSCOPE_IT_POSTGRES_URL must include a plain database name, got {db!r}')
        if not re.fullmatch(r'\w+', db or '') or not db.startswith('erdscope'):
            raise RuntimeError(
                'ERDSCOPE_IT_POSTGRES_URL must name a disposable database whose name '
                f'starts with "erdscope" — it is DROPPED and recreated. Got {db!r}')
        cls.db = db
        host, port, user = u.hostname or '127.0.0.1', u.port or 5432, u.username or 'postgres'
        env = dict(os.environ)
        if u.password:
            env['PGPASSWORD'] = u.password
        base = ['psql', '-X', '-q', '-v', 'ON_ERROR_STOP=1', '-h', host, '-p', str(port), '-U', user]
        subprocess.run(base + ['-d', 'postgres', '-c', f'DROP DATABASE IF EXISTS {db};'],
                        check=True, env=env, capture_output=True, text=True)
        subprocess.run(base + ['-d', 'postgres', '-c', f'CREATE DATABASE {db};'],
                        check=True, env=env, capture_output=True, text=True)
        fixture_sql = (FIXTURE_DIR / 'demo_postgres.sql').read_text(encoding='utf-8')
        subprocess.run(base + ['-d', db], input=fixture_sql, check=True, env=env,
                        capture_output=True, text=True)

    def test_schema_matches_across_driver_and_cli_paths(self):
        with tempfile.TemporaryDirectory() as tmp:
            out_driver = Path(tmp) / 'driver.html'
            out_cli = Path(tmp) / 'cli.html'
            r1 = _run_erd(POSTGRES_URL, out_driver)
            self.assertEqual(r1.returncode, 0, r1.stderr)
            r2 = _run_erd(POSTGRES_URL, out_cli, cli_fallback=True)
            self.assertEqual(r2.returncode, 0, r2.stderr)
            data_driver = _load_data(out_driver)
            data_cli = _load_data(out_cli)
            _assert_no_stray_null_strings(data_cli['tables'])
            self.assertEqual(data_driver, data_cli,
                'psycopg driver path and psql-CLI fallback path must produce '
                'byte-identical IR')
            self._assert_schema(data_driver['tables'])

    def _assert_schema(self, tables):
        self.assertEqual(len(tables), 13)
        self.assertNotIn('active_users', tables)  # VIEW must be excluded

        users = tables['users']
        self.assertEqual(users['comment'], 'ユーザー')
        name_col = next(c for c in users['columns'] if c['name'] == 'name')
        self.assertEqual(name_col['comment'], '氏名')
        status_col = next(c for c in users['columns'] if c['name'] == 'status')
        self.assertEqual(status_col['comment'], '1: active\n2: suspended')

        addresses = tables['addresses']
        self.assertEqual(addresses['comment'], '住所\tタブ入りコメント')

        # shipments.order_id is UNIQUE, so the DB FK promotes to 1:1
        ship_assocs = tables['shipments']['associations']
        self.assertEqual(len(ship_assocs), 1)
        self.assertEqual(ship_assocs[0]['type'], 'has_one')
        self.assertEqual(ship_assocs[0]['target'], 'orders')
        self.assertTrue(ship_assocs[0]['db_fk'])

        # composite primary key
        pc_cols = {c['name']: c for c in tables['product_categories']['columns']}
        self.assertTrue(pc_cols['product_id']['primary'])
        self.assertTrue(pc_cols['category_id']['primary'])

        # composite index
        addr_idx = {ix['name']: ix for ix in addresses['indexes']}
        self.assertEqual(addr_idx['idx_addresses_user_kind']['columns'], ['user_id', 'kind'])

        # expression index (lower(title)) — postgres normalizes it to
        # lower(title::text) in pg_get_indexdef, that's expected
        products_idx = {ix['name']: ix for ix in tables['products']['indexes']}
        expr_idx = products_idx['idx_products_title_lower']
        self.assertEqual(len(expr_idx['columns']), 1)
        self.assertIn('lower(', expr_idx['columns'][0])
        self.assertIn('title', expr_idx['columns'][0])

    def test_schema_query_param_selects_app2(self):
        sep = '&' if '?' in POSTGRES_URL else '?'
        url = f'{POSTGRES_URL}{sep}schema=app2'
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'out.html'
            r = _run_erd(url, out)
            self.assertEqual(r.returncode, 0, r.stderr)
            data = _load_data(out)
            self.assertEqual(set(data['tables']), {'widgets'})

    def test_excel_export_generated(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'out.html'
            xlsx = Path(tmp) / 'out.xlsx'
            r = _run_erd(POSTGRES_URL, out, extra_args=['--excel', str(xlsx)])
            self.assertEqual(r.returncode, 0, r.stderr)
            self.assertTrue(out.exists())
            self.assertTrue(xlsx.exists())
            import openpyxl  # optional dep; installed alongside psycopg in CI
            wb = openpyxl.load_workbook(xlsx)
            self.assertIn('users', wb.sheetnames)


if __name__ == '__main__':
    unittest.main()

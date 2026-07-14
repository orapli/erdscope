"""Unit tests for erd.py

Run from the repository root:
    python3 -m unittest discover -s tests -v

The database is the required source, so most tests exercise the pure
IR-building functions (mysql_ir) and the code-overlay parsers directly.
tests/fixture_app (Rails models), tests/fixture_prisma and
tests/fixture_django cover the --models overlay parsers.
"""
import importlib.util
import json
import os
import re
import subprocess
import sys
import tempfile
import types
import unittest
import zipfile
from pathlib import Path
from types import SimpleNamespace

ROOT = Path(__file__).resolve().parent.parent
FIXTURE = Path(__file__).resolve().parent / 'fixture_app'

spec = importlib.util.spec_from_file_location('erd', ROOT / 'erd.py')
erd = importlib.util.module_from_spec(spec)
spec.loader.exec_module(erd)


# information_schema fixture rows shared across tests
TABLE_ROWS = [
    ('users', 'User accounts'),
    ('posts', ''),
    ('comments', ''),
    ('tags', ''),
    ('posts_tags', ''),
    ('likes', ''),
    ('old_items', ''),
    ('people', ''),
    ('audit_logs', ''),
]
def _col(t, name, dtype, ctype, null='YES', key='', default='', extra='', comment=''):
    return (t, name, dtype, ctype, null, key, default, extra, comment)
COL_ROWS = [
    _col('users', 'id', 'bigint', 'bigint', 'NO', 'PRI', '', 'auto_increment'),
    _col('users', 'email', 'varchar', 'varchar(255)', 'NO', 'UNI',
         comment='Login email address'),
    _col('users', 'bio', 'text', 'text'),
    _col('posts', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
    _col('posts', 'user_id', 'bigint', 'bigint', 'NO', 'MUL'),
    _col('posts', 'title', 'varchar', 'varchar(200)', 'NO'),
    _col('comments', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
    _col('comments', 'post_id', 'bigint', 'bigint', 'NO'),
    _col('comments', 'user_id', 'bigint', 'bigint', 'NO'),
    _col('tags', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
    _col('tags', 'label', 'varchar', 'varchar(50)', 'NO'),
    _col('posts_tags', 'post_id', 'bigint', 'bigint', 'NO'),
    _col('posts_tags', 'tag_id', 'bigint', 'bigint', 'NO'),
    _col('likes', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
    _col('likes', 'post_id', 'bigint', 'bigint', 'NO'),
    _col('likes', 'user_id', 'bigint', 'bigint', 'NO'),
    _col('likes', 'unknown_thing_id', 'bigint', 'bigint'),
    _col('old_items', 'legacy_id', 'bigint', 'bigint', 'NO', 'PRI'),
    _col('people', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
    _col('audit_logs', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
    _col('audit_logs', 'user_id', 'bigint', 'bigint'),
    # view column — no entry in TABLE_ROWS, must be skipped
    _col('v_recent', 'id', 'bigint', 'bigint', 'NO'),
]
FK_ROWS = [
    ('posts', 'user_id', 'users'),
    ('posts_tags', 'post_id', 'posts'),
    ('posts_tags', 'tag_id', 'tags'),
]
INDEX_ROWS = [
    ('users', 'PRIMARY', 0, 1, 'id'),
    ('users', 'index_users_on_email', 0, 1, 'email'),
    ('posts', 'index_posts_on_user_id_and_title', 1, 1, 'user_id'),
    ('posts', 'index_posts_on_user_id_and_title', 1, 2, 'title'),
]

def db_tables():
    return erd.mysql_ir(TABLE_ROWS, COL_ROWS, FK_ROWS, INDEX_ROWS)


# --- new-path helpers -------------------------------------------------------
# The old direct overlay/dedupe/manual-relation functions were retired; the
# live pipeline builds ProviderResult layers and folds them with merge_ir
# (whose Phase B reconcile_db_fks subsumes the old per-table dedupe pass).
# These thin wrappers drive that same path from a DB IR dict so the behavioral
# assertions below carry over unchanged.
def overlay_code(db_ir, mroot, table_map=None):
    """DB IR + framework (Rails/Prisma/Django) layer -> merged IR via merge_ir.
    Reconciliation (the old dedupe pass) is included, since merge_ir runs Phase B."""
    db_layer = erd.make_provider_result('db', 'mysql', db_ir)
    fw = erd.framework_provider(mroot, table_map)
    return erd.merge_ir([db_layer, fw])

def apply_relations(db_ir, relations):
    """DB IR + config relations layer -> merged IR via merge_ir. Validation
    (unknown table/column/target, missing key) happens in
    relations_to_config_layer and raises SystemExit with the same messages."""
    db_layer = erd.make_provider_result('db', 'mysql', db_ir)
    cfg = erd.relations_to_config_layer(relations, db_ir)
    return erd.merge_ir([db_layer, cfg])


class TestInflector(unittest.TestCase):
    def test_pluralize(self):
        cases = {'user': 'users', 'category': 'categories', 'status': 'statuses',
                 'box': 'boxes', 'branch': 'branches', 'leaf': 'leaves',
                 'person': 'people', 'child': 'children'}
        for word, expected in cases.items():
            self.assertEqual(erd.pluralize(word), expected, word)

    def test_to_snake(self):
        self.assertEqual(erd.to_snake('UserStat'), 'user_stat')
        self.assertEqual(erd.to_snake('APIKey'), 'api_key')

    def test_class_to_table(self):
        self.assertEqual(erd.class_to_table('User'), 'users')
        self.assertEqual(erd.class_to_table('Person'), 'people')
        self.assertEqual(erd.class_to_table('Admin::AuditLog'), 'audit_logs')


class TestPyMysqlErrorHandling(unittest.TestCase):
    """mysql_query_rows prefers PyMySQL when it's importable. A connection
    or query failure there must exit cleanly (sys.exit with a one-line
    message) rather than let a raw pymysql traceback through — the CLI
    fallback path already had this, PyMySQL's didn't."""
    def setUp(self):
        class FakeMySQLError(Exception):
            pass
        class FakeErrModule:
            MySQLError = FakeMySQLError
        self.FakeMySQLError = FakeMySQLError
        self.fake_pymysql = types.SimpleNamespace(err=FakeErrModule(), connect=None)
        self.addCleanup(sys.modules.pop, 'pymysql', None)
        sys.modules['pymysql'] = self.fake_pymysql

    def test_connect_failure_exits_cleanly(self):
        def fail_connect(**kw):
            raise self.FakeMySQLError("Access denied for user 'x'@'host'")
        self.fake_pymysql.connect = fail_connect
        with self.assertRaises(SystemExit) as cm:
            erd.mysql_query_rows('mysql://x@127.0.0.1:3306/db', 'SELECT 1')
        self.assertIn('Access denied', str(cm.exception))

    def test_query_failure_exits_cleanly(self):
        class FakeCursor:
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def execute(self, sql):
                raise self.outer.FakeMySQLError('unknown table')
        FakeCursor.outer = self
        class FakeConn:
            def cursor(self): return FakeCursor()
            def close(self): pass
        self.fake_pymysql.connect = lambda **kw: FakeConn()
        with self.assertRaises(SystemExit) as cm:
            erd.mysql_query_rows('mysql://x@127.0.0.1:3306/db', 'SELECT 1')
        self.assertIn('unknown table', str(cm.exception))


class TestUnescapeMysqlField(unittest.TestCase):
    """The mysql-CLI fallback path runs without --raw specifically so
    comments containing tabs/newlines don't corrupt the tab-separated
    output's field count or record boundaries. These test the unescaping
    that makes that work."""
    def test_null_marker(self):
        self.assertEqual(erd._unescape_mysql_field('\\N'), '')

    def test_tab_unescaped(self):
        self.assertEqual(erd._unescape_mysql_field('a\\tb'), 'a\tb')

    def test_newline_unescaped(self):
        self.assertEqual(erd._unescape_mysql_field('line1\\nline2'), 'line1\nline2')

    def test_carriage_return_unescaped(self):
        self.assertEqual(erd._unescape_mysql_field('a\\rb'), 'a\rb')

    def test_escaped_backslash_not_confused_with_following_escape(self):
        # raw bytes: a, \, \, t, b — an escaped literal backslash followed by
        # a literal 't', NOT an escaped tab. Independent chained .replace()
        # calls get this wrong; a single left-to-right pass doesn't.
        self.assertEqual(erd._unescape_mysql_field('a\\\\tb'), 'a\\tb')

    def test_plain_string_unchanged(self):
        self.assertEqual(erd._unescape_mysql_field('plain text'), 'plain text')

    def test_literal_string_null_is_not_the_marker(self):
        # only the exact two-char field \N means SQL NULL — a real comment
        # whose text happens to be the word NULL must survive as data
        self.assertEqual(erd._unescape_mysql_field('NULL'), 'NULL')

    def test_mysql_cli_fallback_survives_tab_and_newline_in_a_field(self):
        # integration-level: mysql_query_rows must not pass --raw (which
        # would leave a real tab/newline in a comment as a literal byte,
        # corrupting split('\t') and splitlines()), and must unescape the
        # \t/\n it gets back instead
        import subprocess as subprocess_mod
        orig_run = subprocess_mod.run
        # multi-line, tab-containing comment plus a NULL column, exactly as
        # mysql --batch (non-raw) would escape them in real TSV output
        fake_stdout = 'users\tid\t\\N\ncomments\tbody\thas a\\ttab and\\nnewline\n'
        seen_cmds = []
        def fake_run(cmd, **kw):
            seen_cmds.append(cmd)
            return subprocess_mod.CompletedProcess(cmd, 0, stdout=fake_stdout, stderr='')
        subprocess_mod.run = fake_run
        self.addCleanup(setattr, subprocess_mod, 'run', orig_run)
        rows = erd.mysql_query_rows('mysql://readonly@127.0.0.1:3306/testdb', 'SELECT 1')
        self.assertNotIn('--raw', seen_cmds[0])
        self.assertEqual(rows, [
            ('users', 'id', ''),
            ('comments', 'body', 'has a\ttab and\nnewline'),
        ])


class TestParseMysqlUrl(unittest.TestCase):
    def test_missing_database_name_exits(self):
        with self.assertRaises(SystemExit) as cm:
            erd.parse_mysql('mysql://readonly@127.0.0.1:3306/')
        self.assertIn('database name', str(cm.exception))

    def test_non_word_database_name_exits(self):
        with self.assertRaises(SystemExit):
            erd.parse_mysql('mysql://readonly@127.0.0.1:3306/db;drop table x')

    def test_valid_url_queries_information_schema_for_db(self):
        seen = []
        orig = erd.mysql_query_rows
        erd.mysql_query_rows = lambda url, sql: (seen.append(sql) or [])
        try:
            erd.parse_mysql('mysql://readonly@127.0.0.1:3306/myapp_production')
        finally:
            erd.mysql_query_rows = orig
        self.assertEqual(len(seen), 4)  # tables, columns, FKs, indexes
        self.assertTrue(all("TABLE_SCHEMA='myapp_production'" in q for q in seen))

    def _parse_with_stubs(self, url, isatty, getpass_return='should-not-be-called'):
        """Run parse_mysql with mysql_query_rows/isatty/getpass stubbed out.
        Stub restoration is registered via addCleanup; MYSQL_PWD is left
        exactly as parse_mysql leaves it so tests can assert on it — callers
        own their own MYSQL_PWD setup/teardown."""
        orig_query = erd.mysql_query_rows
        orig_isatty = sys.stdin.isatty
        orig_getpass = erd.getpass.getpass
        self.addCleanup(lambda: setattr(erd, 'mysql_query_rows', orig_query))
        self.addCleanup(lambda: setattr(sys.stdin, 'isatty', orig_isatty))
        self.addCleanup(lambda: setattr(erd.getpass, 'getpass', orig_getpass))
        erd.mysql_query_rows = lambda url, sql: []
        sys.stdin.isatty = lambda: isatty

        def fake_getpass(prompt=''):
            self._getpass_prompt = prompt
            return getpass_return
        erd.getpass.getpass = fake_getpass
        erd.parse_mysql(url)

    def test_password_prompt_skipped_when_mysql_pwd_already_set(self):
        os.environ['MYSQL_PWD'] = 'preset'
        self.addCleanup(os.environ.pop, 'MYSQL_PWD', None)
        self._parse_with_stubs('mysql://readonly@127.0.0.1:3306/db', isatty=True)
        # getpass would have recorded a prompt if called; confirm it wasn't
        self.assertFalse(hasattr(self, '_getpass_prompt'))
        self.assertEqual(os.environ['MYSQL_PWD'], 'preset')  # left untouched

    def test_password_prompt_skipped_when_not_interactive(self):
        os.environ.pop('MYSQL_PWD', None)
        self._parse_with_stubs('mysql://readonly@127.0.0.1:3306/db', isatty=False)
        self.assertFalse(hasattr(self, '_getpass_prompt'))
        self.assertNotIn('MYSQL_PWD', os.environ)

    def test_password_prompted_when_interactive_and_unset(self):
        os.environ.pop('MYSQL_PWD', None)
        self.addCleanup(os.environ.pop, 'MYSQL_PWD', None)
        self._parse_with_stubs('mysql://readonly@127.0.0.1:3306/db',
                                isatty=True, getpass_return='hunter2')
        self.assertIn('readonly@127.0.0.1', self._getpass_prompt)
        self.assertEqual(os.environ.get('MYSQL_PWD'), 'hunter2')

    def test_password_prompt_skipped_when_url_has_explicit_empty_password(self):
        # mysql://user:@host/db — an explicit empty password should be
        # respected as "no password", not treated as "unset, please prompt"
        os.environ.pop('MYSQL_PWD', None)
        self._parse_with_stubs('mysql://readonly:@127.0.0.1:3306/db', isatty=True)
        self.assertFalse(hasattr(self, '_getpass_prompt'))


class TestMySQLIr(unittest.TestCase):
    def setUp(self):
        self.tables = db_tables()

    def test_views_are_skipped(self):
        self.assertNotIn('v_recent', self.tables)
        self.assertEqual(len(self.tables), len(TABLE_ROWS))

    def test_table_comment(self):
        self.assertEqual(self.tables['users']['comment'], 'User accounts')
        self.assertNotIn('comment', self.tables['posts'])

    def test_types_nullable_pk(self):
        cols = {c['name']: c for c in self.tables['users']['columns']}
        self.assertEqual(cols['email']['type'], 'string')
        self.assertEqual(cols['email']['sql_type'], 'varchar(255)')
        self.assertFalse(cols['email']['nullable'])
        self.assertTrue(cols['bio']['nullable'])
        self.assertTrue(cols['id']['primary'])
        self.assertEqual(cols['id']['extra'], 'auto_increment')
        self.assertEqual(self.tables['users']['primary_key'], 'id')
        self.assertEqual(self.tables['old_items']['primary_key'], 'legacy_id')

    def test_column_comment(self):
        cols = {c['name']: c for c in self.tables['users']['columns']}
        self.assertEqual(cols['email']['comment'], 'Login email address')
        self.assertNotIn('comment', cols['bio'])

    def test_db_fk_association(self):
        a = self.tables['posts']['associations'][0]
        self.assertEqual((a['type'], a['name'], a['target'], a['foreign_key']),
                         ('belongs_to', 'user', 'users', 'user_id'))
        self.assertTrue(a['db_fk'])

    def test_indexes(self):
        idx = {i['name']: i for i in self.tables['users']['indexes']}
        self.assertIn('PRIMARY', idx)
        self.assertTrue(idx['index_users_on_email']['unique'])
        multi = self.tables['posts']['indexes'][0]
        self.assertEqual(multi['columns'], ['user_id', 'title'])  # SEQ order
        self.assertFalse(multi['unique'])

    def test_db_fk_under_unique_index_is_1to1(self):
        # a FK column alone under a UNIQUE index can't repeat — real 1:1,
        # not the default many:1 a bare FK column implies
        table_rows = [('accounts', ''), ('profiles', '')]
        col_rows = [
            _col('accounts', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('profiles', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('profiles', 'account_id', 'bigint', 'bigint', 'NO', 'UNI'),
        ]
        fk_rows = [('profiles', 'account_id', 'accounts')]
        index_rows = [('profiles', 'uk_profiles_account_id', 0, 1, 'account_id')]
        tables = erd.mysql_ir(table_rows, col_rows, fk_rows, index_rows)
        a = tables['profiles']['associations'][0]
        self.assertEqual(a['type'], 'has_one')
        self.assertEqual(a['foreign_key'], 'account_id')


class TestOverlayAndInference(unittest.TestCase):
    def setUp(self):
        self.tables = db_tables()

    def test_rails_overlay_merges_associations(self):
        self.assertEqual(erd.detect_code_source(FIXTURE), 'rails')
        self.tables = overlay_code(self.tables, FIXTURE)
        names = {a['name'] for a in self.tables['users']['associations']}
        self.assertIn('posts', names)
        self.assertIn('commented_posts', names)  # has_many :through

    def test_belongs_to_backfills_conventional_foreign_key(self):
        # comment.rb: `belongs_to :post` / `belongs_to :user` — neither gives
        # an explicit foreign_key: option, so it must default to Rails
        # convention (<name>_id) rather than being left unset
        self.tables = overlay_code(self.tables, FIXTURE)
        assocs = {a['name']: a for a in self.tables['comments']['associations']}
        self.assertEqual(assocs['post']['foreign_key'], 'post_id')
        self.assertEqual(assocs['user']['foreign_key'], 'user_id')
        # an explicit foreign_key: option is still honored, not overridden
        post_assocs = {a['name']: a for a in self.tables['posts']['associations']}
        self.assertEqual(post_assocs['author']['foreign_key'], 'user_id')

    def test_dedupe_after_overlay(self):
        # merge_ir's Phase B reconcile_db_fks does what the old dedupe pass did:
        # posts.user_id DB FK is covered by the explicit belongs_to and dropped,
        # while posts_tags FKs stay (no model declares that pair). The behavioral
        # signal (the pair got reconciled) is the resulting db_fk state below.
        # the merged IR carries structured provenance, not the legacy db_fk flag
        merged = overlay_code(self.tables, FIXTURE)
        self.assertFalse(any(a.get('provenance') == 'db_fk'
                             for a in merged['posts']['associations']))
        self.assertTrue(any(a.get('provenance') == 'db_fk'
                            for a in merged['posts_tags']['associations']))

    def test_model_without_db_table_flagged(self):
        self.tables = overlay_code(self.tables, FIXTURE)
        self.assertIn('webhooks', self.tables)
        self.assertTrue(self.tables['webhooks']['schema_missing'])

    def test_table_name_resolved_through_included_concern(self):
        # widget.rb declares no self.table_name itself — it's set inside
        # `included do` in the HasCrmTable concern it includes. Without
        # following the include, this would fall back to class_to_table
        # ('Widget' -> 'widgets'), which is wrong.
        self.tables = overlay_code(self.tables, FIXTURE)
        self.assertIn('crm_widgets', self.tables)
        self.assertNotIn('widgets', self.tables)

    def test_table_map_overrides_unresolvable_table_name(self):
        # gizmo.rb includes a concern that lives in a gem, not this app —
        # genuinely unresolvable by static analysis, so it falls back to
        # class_to_table('Gizmo') = 'gizmos', which is wrong for this model.
        self.tables = overlay_code(self.tables, FIXTURE)
        self.assertIn('gizmos', self.tables)  # the (wrong) naive guess, unmapped

    def test_table_map_corrects_it(self):
        self.tables = overlay_code(self.tables, FIXTURE, table_map={'Gizmo': 'crm_gizmos'})
        self.assertIn('crm_gizmos', self.tables)
        self.assertNotIn('gizmos', self.tables)
        # the override wins even over a *correctly* resolved table_name
        erd2 = overlay_code(db_tables(), FIXTURE, table_map={'Widget': 'totally_different'})
        self.assertIn('totally_different', erd2)
        self.assertNotIn('crm_widgets', erd2)

    def test_model_on_custom_base_class_is_not_silently_dropped(self):
        # custom_base_widget.rb: `class CustomBaseWidget < BaseRecord`, and
        # base_record.rb: `class BaseRecord < ApplicationRecord`. Only a
        # literal `< ApplicationRecord`/`< ActiveRecord::Base` used to be
        # recognized, so a model built on a shared custom base class (common
        # in mature Rails apps) was dropped with no warning at all — its
        # associations (belongs_to :user here) vanished silently.
        self.tables = overlay_code(self.tables, FIXTURE)
        self.assertIn('custom_base_widgets', self.tables)
        names = {a['name'] for a in self.tables['custom_base_widgets']['associations']}
        self.assertIn('user', names)
        # the abstract base class itself must not become a bogus table
        self.assertNotIn('base_records', self.tables)

    def test_abstract_class_detection_scoped_to_its_own_class_not_whole_file(self):
        # multi_class_file.rb declares SharedBase (abstract) AND Gadget
        # (concrete, using SharedBase) in the SAME file. The abstract-class
        # regex must be scoped to SharedBase's own body — if it were read
        # from the whole file instead, Gadget would incorrectly inherit
        # "abstract" too and get silently dropped, the exact failure this
        # base-class resolution was built to fix in the first place.
        self.tables = overlay_code(self.tables, FIXTURE)
        self.assertIn('gadgets', self.tables)
        names = {a['name'] for a in self.tables['gadgets']['associations']}
        self.assertIn('user', names)
        self.assertNotIn('shared_bases', self.tables)

    def test_sti_subclass_shares_parent_table_not_a_phantom_one(self):
        # admin.rb: `class Admin < User` — User is a *concrete* model
        # (posts.rb's User has real associations, no abstract_class), so
        # this is Rails single-table inheritance: Admin shares users' table
        # rather than getting its own. Admin's own association
        # (belongs_to :department) must land on `users`, not a phantom
        # `admins` table that doesn't exist in the real schema.
        self.tables = overlay_code(self.tables, FIXTURE)
        self.assertNotIn('admins', self.tables)
        names = {a['name'] for a in self.tables['users']['associations']}
        self.assertIn('department', names)

    def test_table_map_on_an_sti_subclass_still_wins(self):
        # table_map is the documented "static analysis genuinely can't
        # reach this" escape hatch — an explicit override on the STI
        # subclass itself must still take precedence over sti_root's
        # automatic table-sharing, not be silently ignored by it
        self.tables = overlay_code(self.tables, FIXTURE, table_map={'Admin': 'admin_accounts'})
        self.assertIn('admin_accounts', self.tables)
        names = {a['name'] for a in self.tables['admin_accounts']['associations']}
        self.assertIn('department', names)
        # and it must not also have leaked into users
        self.assertNotIn('department',
                         {a['name'] for a in self.tables['users']['associations']})

    def test_associations_pointing_at_a_renamed_table_use_the_real_name(self):
        # project.rb: `class Project < ApplicationRecord; self.table_name =
        # 'aaa_projects'; end`. task.rb references it two ways: implicit
        # `belongs_to :project` (no class_name:) and explicit `belongs_to
        # :owner, class_name: 'Project'`. Both used to resolve their target
        # via the naive class_to_table('Project') = 'projects' — a table
        # that doesn't exist, since the real one is aaa_projects — so the
        # right-pane link pointed nowhere and was unclickable.
        self.tables = overlay_code(self.tables, FIXTURE)
        self.assertIn('aaa_projects', self.tables)
        self.assertNotIn('projects', self.tables)
        targets = {a['name']: a['target'] for a in self.tables['tasks']['associations']}
        self.assertEqual(targets['project'], 'aaa_projects')
        self.assertEqual(targets['owner'], 'aaa_projects')

    def test_commented_out_table_name_does_not_win(self):
        # commented_table_name.rb has `# self.table_name = 'should_not_be_used'`
        # — the self.table_name regex must run on comment-stripped source,
        # or a commented-out assignment silently overrides the real
        # (class_to_table-derived) table name
        self.tables = overlay_code(self.tables, FIXTURE)
        self.assertIn('commented_table_names', self.tables)
        self.assertNotIn('should_not_be_used', self.tables)

    def test_dedupe_upgrades_lone_belongs_to_when_db_says_1to1(self):
        # simulates an incomplete Rails declaration: `belongs_to :account`
        # with no matching `has_one` on the other side, so the code alone
        # never asserts cardinality — but the DB has a unique index on the
        # FK column, so dedupe should promote the belongs_to to has_one
        # rather than just deleting the DB FK and losing that signal
        table_rows = [('accounts', ''), ('profiles', '')]
        col_rows = [
            _col('accounts', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('profiles', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('profiles', 'account_id', 'bigint', 'bigint', 'NO', 'UNI'),
        ]
        fk_rows = [('profiles', 'account_id', 'accounts')]
        index_rows = [('profiles', 'uk_profiles_account_id', 0, 1, 'account_id')]
        tables = erd.mysql_ir(table_rows, col_rows, fk_rows, index_rows)
        self.assertEqual(tables['profiles']['associations'][0]['type'], 'has_one')

        # an incomplete Rails-style overlay: only the belongs_to side declared
        tables['profiles']['associations'].append(
            {'type': 'belongs_to', 'name': 'account', 'target': 'accounts',
             'foreign_key': 'account_id'})

        removed = erd.reconcile_db_fks(tables)
        self.assertEqual(removed, 1)
        assocs = tables['profiles']['associations']
        self.assertEqual(len(assocs), 1, 'the DB FK should be dropped, not duplicated')
        self.assertEqual(assocs[0]['type'], 'has_one')
        self.assertNotIn('db_fk', assocs[0])  # still the explicit (code-declared) entry

    def test_dedupe_leaves_explicit_cardinality_alone(self):
        # when the code *does* declare cardinality (has_many here), dedupe
        # must not second-guess it even if the DB FK resolved to has_one
        table_rows = [('accounts', ''), ('profiles', '')]
        col_rows = [
            _col('accounts', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('profiles', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('profiles', 'account_id', 'bigint', 'bigint', 'NO', 'UNI'),
        ]
        fk_rows = [('profiles', 'account_id', 'accounts')]
        index_rows = [('profiles', 'uk_profiles_account_id', 0, 1, 'account_id')]
        tables = erd.mysql_ir(table_rows, col_rows, fk_rows, index_rows)
        tables['profiles']['associations'].append(
            {'type': 'belongs_to', 'name': 'account', 'target': 'accounts',
             'foreign_key': 'account_id'})
        tables['accounts']['associations'].append(
            {'type': 'has_many', 'name': 'profiles', 'target': 'profiles'})

        erd.reconcile_db_fks(tables)
        types = {a['type'] for a in tables['profiles']['associations']}
        self.assertEqual(types, {'belongs_to'})  # left as-is, not upgraded

    def test_dedupe_is_column_aware_not_just_pair_aware(self):
        # posts has TWO distinct FK columns pointing at users
        # (author_id, editor_id) — an explicit belongs_to naming author_id
        # must only dedupe the author_id DB FK, not also swallow editor_id
        table_rows = [('users', ''), ('posts', '')]
        col_rows = [
            _col('users', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('posts', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('posts', 'author_id', 'bigint', 'bigint'),
            _col('posts', 'editor_id', 'bigint', 'bigint'),
        ]
        fk_rows = [('posts', 'author_id', 'users'), ('posts', 'editor_id', 'users')]
        tables = erd.mysql_ir(table_rows, col_rows, fk_rows, [])
        tables['posts']['associations'].append(
            {'type': 'belongs_to', 'name': 'author', 'target': 'users',
             'foreign_key': 'author_id'})

        removed = erd.reconcile_db_fks(tables)
        self.assertEqual(removed, 1)  # only the author_id DB FK
        fks = {a['foreign_key'] for a in tables['posts']['associations']}
        self.assertEqual(fks, {'author_id', 'editor_id'})  # editor_id DB FK survives
        editor = next(a for a in tables['posts']['associations']
                     if a['foreign_key'] == 'editor_id')
        self.assertTrue(editor.get('db_fk'))

    def test_inference_on_constraintless_fk(self):
        added = erd.infer_fk_associations(self.tables)
        targets = {a['target'] for a in self.tables['likes']['associations']}
        self.assertEqual(targets, {'posts', 'users'})
        self.assertTrue(all(a['provenance'] == 'inferred'
                            for a in self.tables['likes']['associations']))
        self.assertGreaterEqual(added, 2)

    def test_inference_skips_covered_pairs(self):
        erd.infer_fk_associations(self.tables)
        # posts.user_id has a DB FK — no duplicate inferred edge
        assocs = [a for a in self.tables['posts']['associations']
                  if a['target'] == 'users']
        self.assertEqual(len(assocs), 1)

    def test_inference_falls_back_to_singular_table_name(self):
        # Rails-style pluralized tables are tried first, but schemas that
        # don't pluralize (common outside Rails, e.g. Prisma defaults) should
        # still resolve — 'author_id' -> 'author' when 'authors' doesn't exist
        table_rows = [('posts', ''), ('author', '')]  # singular, deliberately
        col_rows = [
            _col('posts', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('posts', 'author_id', 'bigint', 'bigint'),
            _col('author', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
        ]
        tables = erd.mysql_ir(table_rows, col_rows, [], [])
        added = erd.infer_fk_associations(tables)
        self.assertEqual(added, 1)
        a = tables['posts']['associations'][0]
        self.assertEqual((a['target'], a['provenance']), ('author', 'inferred'))

    def test_inference_detects_1to1_via_unique_index(self):
        table_rows = [('users', ''), ('profiles', '')]
        col_rows = [
            _col('users', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('profiles', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('profiles', 'user_id', 'bigint', 'bigint', 'NO', 'UNI'),
        ]
        index_rows = [('profiles', 'uk_profiles_user_id', 0, 1, 'user_id')]
        tables = erd.mysql_ir(table_rows, col_rows, [], index_rows)
        erd.infer_fk_associations(tables)
        a = tables['profiles']['associations'][0]
        self.assertEqual(a['type'], 'has_one')


class TestManualRelations(unittest.TestCase):
    def setUp(self):
        self.tables = db_tables()

    def test_declares_belongs_to_by_default(self):
        merged = apply_relations(self.tables, [
            {'table': 'comments', 'column': 'post_id', 'references': 'posts'},
        ])
        a = merged['comments']['associations'][0]
        self.assertEqual((a['type'], a['name'], a['target'], a['foreign_key']),
                         ('belongs_to', 'post', 'posts', 'post_id'))
        self.assertEqual(a['provenance'], 'manual')  # merge_ir forces config-kind to manual

    def test_one_to_one_flag_forces_has_one(self):
        merged = apply_relations(self.tables, [
            {'table': 'comments', 'column': 'post_id', 'references': 'posts',
             'one_to_one': True, 'name': 'thread'},
        ])
        a = merged['comments']['associations'][0]
        self.assertEqual((a['type'], a['name']), ('has_one', 'thread'))

    def test_auto_detects_1to1_via_unique_index_like_other_sources(self):
        idx = {'name': 'uk_comments_post_id', 'columns': ['post_id'], 'unique': True}
        self.tables['comments']['indexes'].append(idx)
        merged = apply_relations(self.tables, [
            {'table': 'comments', 'column': 'post_id', 'references': 'posts'},
        ])
        self.assertEqual(merged['comments']['associations'][0]['type'], 'has_one')

    def test_works_without_any_models_overlay(self):
        # the whole point: a complete relation graph from config alone,
        # usable before any application code exists
        merged = apply_relations(self.tables, [
            {'table': 'likes', 'column': 'unknown_thing_id', 'references': 'users'},
        ])
        self.assertIn('users', {a['target'] for a in merged['likes']['associations']})

    def test_unknown_table_exits(self):
        with self.assertRaises(SystemExit) as cm:
            apply_relations(self.tables, [
                {'table': 'nope', 'column': 'x_id', 'references': 'users'},
            ])
        self.assertIn('nope', str(cm.exception))

    def test_unknown_column_exits(self):
        with self.assertRaises(SystemExit) as cm:
            apply_relations(self.tables, [
                {'table': 'comments', 'column': 'nope_id', 'references': 'posts'},
            ])
        self.assertIn('nope_id', str(cm.exception))

    def test_unknown_target_exits(self):
        with self.assertRaises(SystemExit) as cm:
            apply_relations(self.tables, [
                {'table': 'comments', 'column': 'post_id', 'references': 'nope'},
            ])
        self.assertIn('nope', str(cm.exception))

    def test_missing_key_exits(self):
        with self.assertRaises(SystemExit) as cm:
            apply_relations(self.tables, [{'table': 'comments', 'column': 'post_id'}])
        self.assertIn('references', str(cm.exception))

    def test_plain_db_fk_does_not_block_a_manual_relation(self):
        # posts.user_id already has a real DB FK to users, but a plain DB FK
        # isn't "explicit" — the config relation takes precedence over it
        # (merge_ir Phase B reconciles the pair), producing the manual edge.
        merged = apply_relations(self.tables, [
            {'table': 'posts', 'column': 'user_id', 'references': 'users', 'name': 'owner'},
        ])
        owner = [a for a in merged['posts']['associations'] if a['name'] == 'owner']
        self.assertEqual(len(owner), 1)
        self.assertEqual(owner[0]['provenance'], 'manual')

    def test_config_relation_overrides_existing_explicit_association(self):
        # P0-3 / §12 change: the OLD path SKIPPED a relation already covered by
        # an explicit association (added == 0). The NEW path OVERRIDES it — a
        # config relation of the same identity wins via merge_ir's Phase A and
        # is re-authored as config/manual. Here a code-declared belongs_to
        # :author (user_id) is overridden by a config relation naming the same
        # edge, yielding a single, config-authored association (not a skip).
        self.tables['posts']['associations'].append(
            {'type': 'belongs_to', 'name': 'author', 'target': 'users',
             'foreign_key': 'user_id'})
        merged = apply_relations(self.tables, [
            {'table': 'posts', 'column': 'user_id', 'references': 'users', 'name': 'author'},
        ])
        matching = [a for a in merged['posts']['associations']
                    if a.get('foreign_key') == 'user_id' and a['target'] == 'users'
                    and a['name'] == 'author']
        self.assertEqual(len(matching), 1)      # merged, not duplicated
        self.assertEqual(matching[0]['provenance'], 'manual')  # config authorship wins

    def test_takes_precedence_over_db_fk_via_reconcile(self):
        # config relation wins over a real (but differently-named) DB FK the
        # same way a code-declared association would: merge_ir Phase B drops the
        # covered DB FK, leaving only the manual edge.
        table_rows = [('a', ''), ('b', '')]
        col_rows = [
            _col('a', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('b', 'id', 'bigint', 'bigint', 'NO', 'PRI'),
            _col('b', 'weird_ref', 'bigint', 'bigint'),
        ]
        fk_rows = [('b', 'weird_ref', 'a')]
        tables = erd.mysql_ir(table_rows, col_rows, fk_rows, [])
        merged = apply_relations(tables, [
            {'table': 'b', 'column': 'weird_ref', 'references': 'a', 'name': 'owner'},
        ])
        assocs = merged['b']['associations']
        self.assertEqual(len(assocs), 1)
        self.assertEqual(assocs[0]['provenance'], 'manual')
        self.assertNotIn('db_fk', assocs[0])  # no legacy booleans on the merged IR


class TestFkColumns(unittest.TestCase):
    """fk_columns (computed in _finish(), the single source of truth the FK
    badge and PK/FK column view read) must only ever contain columns backed
    by a real association — never a bare *_id name match. --infer-fk is the
    only thing that can widen it, and only for names that also resolve to a
    real table (comments.post_id/user_id do; likes.unknown_thing_id never
    can, since no `unknown_things` table exists)."""
    def _args(self, **kw):
        base = dict(output='', models=None, excel=None, max_rows=15,
                    only=None, exclude=None, infer_fk=False)
        base.update(kw)
        return SimpleNamespace(**base)

    def test_default_excludes_unbacked_id_columns(self):
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            erd._finish(tables, self._args(output=str(Path(tmp) / 'out.html')), 'testdb')
        self.assertNotIn('post_id', tables['comments']['fk_columns'])
        self.assertNotIn('user_id', tables['comments']['fk_columns'])
        # posts.user_id has a real DB FK constraint — always included
        self.assertIn('user_id', tables['posts']['fk_columns'])

    def test_infer_fk_widens_it_to_matching_tables_only(self):
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            erd._finish(tables, self._args(output=str(Path(tmp) / 'out.html'), infer_fk=True),
                        'testdb')
        self.assertIn('post_id', tables['comments']['fk_columns'])
        self.assertIn('user_id', tables['comments']['fk_columns'])
        # no `unknown_things` table exists — never inferred, flag or not
        self.assertNotIn('unknown_thing_id', tables['likes']['fk_columns'])


class TestParsePrisma(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tables = erd.parse_prisma(
            Path(__file__).resolve().parent / 'fixture_prisma' / 'schema.prisma')

    def test_models_found_with_map(self):
        self.assertEqual(set(self.tables), {'users', 'Profile', 'Post', 'Tag'})

    def test_relations(self):
        assocs = {a['name']: a for a in self.tables['Post']['associations']}
        self.assertEqual(assocs['author']['type'], 'belongs_to')
        self.assertEqual(assocs['author']['foreign_key'], 'authorId')
        self.assertEqual(assocs['tags']['type'], 'has_and_belongs_to_many')

    def test_at_unique_fk_field_is_1to1_not_manyto1(self):
        # Profile.userId is `Int @unique` — each value can appear at most
        # once, so it's a real 1:1, not the default many:1 a bare FK implies
        assocs = {a['name']: a for a in self.tables['Profile']['associations']}
        self.assertEqual(assocs['user']['type'], 'has_one')
        self.assertEqual(assocs['user']['foreign_key'], 'userId')
        # Post.authorId has no @unique — stays the default many:1
        post_assocs = {a['name']: a for a in self.tables['Post']['associations']}
        self.assertEqual(post_assocs['author']['type'], 'belongs_to')

    def test_prisma_as_overlay(self):
        # prisma overlaid on a DB IR via merge_ir: a matching table gains the
        # prisma associations, and a prisma-only model surfaces as its own
        # table. (Unlike the retired association-only overlay, the prisma
        # provider retains columns, so a prisma-only model is a real table
        # rather than a schema_missing shell.)
        base = {'users': {'columns': [], 'associations': [],
                          'indexes': [], 'primary_key': 'id'}}
        prisma_dir = Path(__file__).resolve().parent / 'fixture_prisma'
        self.assertEqual(erd.detect_code_source(prisma_dir), 'prisma')
        merged = erd.merge_ir([
            erd.make_provider_result('db', 'mysql', base),
            erd.framework_provider(prisma_dir),
        ])
        self.assertTrue(any(a['name'] == 'posts'
                            for a in merged['users']['associations']))
        self.assertIn('Post', merged)


class TestParseDjango(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tables = erd.parse_django(
            Path(__file__).resolve().parent / 'fixture_django')

    def test_tables_found(self):
        self.assertEqual(set(self.tables),
                         {'blog_author', 'blog_entries', 'blog_posttag',
                          'blog_tag', 'shop_product'})

    def test_relations(self):
        assocs = {a['name']: a for a in self.tables['blog_entries']['associations']}
        self.assertEqual(assocs['author']['target'], 'blog_author')
        self.assertEqual(assocs['parent']['target'], 'blog_entries')  # self
        self.assertEqual(assocs['tags']['through'], 'blog_posttag')

    def test_detect_code_source(self):
        base = Path(__file__).resolve().parent
        self.assertEqual(erd.detect_code_source(base / 'fixture_django'), 'django')
        self.assertEqual(erd.detect_code_source(base / 'fixture_prisma'), 'prisma')
        self.assertEqual(erd.detect_code_source(FIXTURE), 'rails')


class TestGeneration(unittest.TestCase):
    def _args(self, **kw):
        base = dict(output='', models=None, excel=None, max_rows=15,
                    only=None, exclude=None, infer_fk=False)
        base.update(kw)
        return SimpleNamespace(**base)

    def test_html_generation(self):
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'out.html'
            erd._finish(tables, self._args(output=str(out)), 'testdb')
            html = out.read_text(encoding='utf-8')
        for ph in ('__DATA_JSON__', '__MAX_ROWS__', '__TITLE__'):
            self.assertNotIn(ph, html)
        self.assertIn('<title>testdb — ERD</title>', html)
        m = re.search(r'^const DATA = (.+);$', html, re.MULTILINE)
        data = json.loads(m.group(1))
        self.assertIn('users', data['tables'])
        self.assertEqual(data['tables']['users']['comment'], 'User accounts')
        self.assertTrue(data['tables']['users']['indexes'])

    def test_comment_with_script_close_tag_does_not_break_the_page(self):
        # a table/column comment is free-text from the database — one
        # containing a literal "</script>" must not be able to prematurely
        # close the embedded <script> block and blank the whole page
        tables = db_tables()
        tables['users']['comment'] = 'nested </script><script>alert(1)</script> payload'
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'out.html'
            erd._finish(tables, self._args(output=str(out)), 'testdb')
            html = out.read_text(encoding='utf-8')
        # the literal closing sequence must never appear unescaped in the
        # output — only the escaped <\/script> form is safe to embed
        self.assertNotIn('</script><script>alert', html)
        # exactly one real </script> (the template's own, closing the block)
        self.assertEqual(html.count('</script>'), 1)
        m = re.search(r'^const DATA = (.+);$', html, re.MULTILINE)
        data = json.loads(m.group(1))
        self.assertIn('</script>', data['tables']['users']['comment'])  # round-trips intact

    def test_comment_containing_template_placeholder_text_is_not_corrupted(self):
        # __TITLE__/__MAX_ROWS__ substitution must happen before DATA_JSON
        # is spliced in, or a comment containing that literal text would
        # get silently rewritten by the later .replace() calls
        tables = db_tables()
        tables['users']['comment'] = 'contains the literal text __TITLE__ and __MAX_ROWS__'
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'out.html'
            erd._finish(tables, self._args(output=str(out)), 'testdb')
            html = out.read_text(encoding='utf-8')
        data = json.loads(re.search(r'^const DATA = (.+);$', html, re.MULTILINE).group(1))
        self.assertEqual(data['tables']['users']['comment'],
                         'contains the literal text __TITLE__ and __MAX_ROWS__')

    def test_only_exclude(self):
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'out.html'
            erd._finish(tables, self._args(output=str(out), only=['post*,tags'],
                                           exclude=['posts_tags']), 'testdb')
            html = out.read_text(encoding='utf-8')
        data = json.loads(re.search(r'^const DATA = (.+);$', html,
                                    re.MULTILINE).group(1))
        self.assertEqual(set(data['tables']), {'posts', 'tags'})

    def test_cli_requires_database_url(self):
        # a bare path (not a URL) as the positional arg has no recognized scheme
        res = subprocess.run([sys.executable, str(ROOT / 'erd.py'),
                              str(FIXTURE)], capture_output=True, text=True)
        self.assertNotEqual(res.returncode, 0)
        self.assertIn('unrecognized database URL scheme', res.stderr)


class TestExcel(unittest.TestCase):
    def test_workbook_structure(self):
        tables = overlay_code(db_tables(), FIXTURE)
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'defs.xlsx'
            erd.write_excel(tables, out, 'testdb')
            with zipfile.ZipFile(out) as z:
                names = z.namelist()
                self.assertIn('xl/workbook.xml', names)
                sheets = [n for n in names if n.startswith('xl/worksheets/')]
                # overview + one per table
                self.assertEqual(len(sheets), 1 + len(tables))
                wb = z.read('xl/workbook.xml').decode()
                self.assertIn('name="Tables"', wb)
                self.assertIn('name="users"', wb)
                users_sheet = None
                for i in range(len(sheets)):
                    xml = z.read(f'xl/worksheets/sheet{i+1}.xml').decode()
                    if '>Login email address<' in xml:
                        users_sheet = xml
                self.assertIsNotNone(users_sheet, 'column comment in a sheet')
                self.assertIn('varchar(255)', users_sheet)
                self.assertIn('index_users_on_email', users_sheet)
                self.assertIn('>UNIQUE<', users_sheet)
                overview = z.read('xl/worksheets/sheet1.xml').decode()
                self.assertIn('User accounts', overview)
                self.assertIn('<hyperlink', overview)

    def test_sheet_name_sanitizing(self):
        used = set()
        self.assertEqual(erd._sheet_name('a/b:c*d', used), 'a_b_c_d')
        long = erd._sheet_name('x' * 40, used)
        self.assertLessEqual(len(long), 31)
        dup = erd._sheet_name('a/b:c*d', used)
        self.assertNotEqual(dup, 'a_b_c_d')

    def test_openpyxl_roundtrip_if_available(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest('openpyxl not installed')
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'defs.xlsx'
            erd.write_excel(tables, out, 'testdb')
            wb = openpyxl.load_workbook(out)
            self.assertEqual(wb.sheetnames[0], 'Tables')
            self.assertIn('users', wb.sheetnames)

    def test_sheet_xml_keeps_a_styled_cell_even_when_its_value_is_empty(self):
        # regression: _sheet_xml used to `continue` on any empty value
        # before even checking for a style, silently dropping the cell —
        # invisible while only header cells carried styles, but once every
        # data cell carries a border/zebra-fill role, a blank Default/
        # Extra/Comment column (extremely common) punched a border-less
        # hole in an otherwise fully-gridded table
        xml = erd._sheet_xml([[('filled', 3), ('', 3), (None, 3), ('', 0)]])
        self.assertIn('<c r="A1" s="3" t="inlineStr">', xml)
        self.assertIn('<c r="B1" s="3"/>', xml, 'an empty value with a style must still emit a styled cell')
        self.assertIn('<c r="C1" s="3"/>', xml, 'None is exactly as empty as "" for this purpose')
        self.assertNotIn('r="D1"', xml, 'an empty value with no style (role 0) is still skipped, as before')

    def test_default_styling_has_title_header_and_zebra_stripes(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest('openpyxl not installed')
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'defs.xlsx'
            erd.write_excel(tables, out, 'testdb')
            ws = openpyxl.load_workbook(out)['Tables']
            self.assertTrue(ws['A1'].font.b)
            self.assertEqual(ws['A1'].font.sz, 14)
            self.assertIsNone(ws['A1'].fill.fgColor.rgb if ws['A1'].fill.patternType else None)
            self.assertTrue(ws['A3'].font.b)  # header row ('#','Table',...)
            self.assertEqual(ws['A3'].font.color.rgb, 'FFFFFFFF')
            self.assertEqual(ws['A3'].fill.fgColor.rgb, 'FF1E293B')
            # first two data rows must alternate (zebra stripe)
            row0_fill = ws['A4'].fill.fgColor.rgb if ws['A4'].fill.patternType == 'solid' else None
            row1_fill = ws['A5'].fill.fgColor.rgb if ws['A5'].fill.patternType == 'solid' else None
            self.assertNotEqual(row0_fill, row1_fill)
            self.assertEqual(row1_fill, 'FFF8FAFC')
            self.assertEqual(ws['A4'].border.left.style, 'thin')

    def _build_template(self, path, styled_cells=(1, 2, 3, 4, 5)):
        """A minimal .xlsx with distinct styling on whichever of A1:A5 are
        listed in styled_cells, via openpyxl (a real editor's OOXML shape,
        not erd.py's own writer) — the closest a unit test gets to Fable's
        recommendation of a fixture "saved by real Excel," without
        checking in a binary file."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Styles'
        thin = Side(style='thin', color='FF00FF00')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        specs = {
            1: (Font(bold=True, size=20, color='FFFF00FF'), None, None),
            2: (Font(bold=True, color='FFFFFFFF'), PatternFill('solid', fgColor='FF123456'), None),
            3: (None, None, border),
            4: (None, PatternFill('solid', fgColor='FFEEEEEE'), border),
            5: (Font(bold=True, italic=True), None, None),
        }
        for row in styled_cells:
            font, fill, brd = specs[row]
            cell = ws.cell(row=row, column=1, value=f'role{row}')
            if font: cell.font = font
            if fill: cell.fill = fill
            if brd: cell.border = brd
        wb.save(path)

    def test_excel_template_overrides_the_default_styling(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest('openpyxl not installed')
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            template = Path(tmp) / 'template.xlsx'
            self._build_template(template)
            out = Path(tmp) / 'defs.xlsx'
            erd.write_excel(tables, out, 'testdb', template_path=str(template))
            ws = openpyxl.load_workbook(out)['Tables']
            self.assertEqual(ws['A1'].font.sz, 20)
            self.assertEqual(ws['A1'].font.color.rgb, 'FFFF00FF')
            self.assertEqual(ws['A3'].fill.fgColor.rgb, 'FF123456')
            self.assertEqual(ws['A4'].border.left.color.rgb, 'FF00FF00')
            self.assertEqual(ws['A5'].fill.fgColor.rgb, 'FFEEEEEE')  # data-alt zebra row

    def test_excel_template_falls_back_role_by_role_for_missing_cells(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest('openpyxl not installed')
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            template = Path(tmp) / 'partial_template.xlsx'
            self._build_template(template, styled_cells=(1,))  # only the title role is styled
            out = Path(tmp) / 'defs.xlsx'
            erd.write_excel(tables, out, 'testdb', template_path=str(template))
            ws = openpyxl.load_workbook(out)['Tables']
            self.assertEqual(ws['A1'].font.sz, 20)  # from the template
            self.assertEqual(ws['A3'].fill.fgColor.rgb, 'FF1E293B')  # built-in header fallback

    def test_bad_excel_template_path_exits_with_a_clear_error(self):
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / 'defs.xlsx'
            with self.assertRaises(SystemExit) as cm:
                erd.write_excel(tables, out, 'testdb', template_path='/no/such/template.xlsx')
            self.assertIn('excel-template', str(cm.exception))

    def test_not_actually_an_xlsx_exits_with_a_clear_error(self):
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            fake = Path(tmp) / 'not_a_workbook.xlsx'
            fake.write_text('this is plain text, not a zip')
            out = Path(tmp) / 'defs.xlsx'
            with self.assertRaises(SystemExit) as cm:
                erd.write_excel(tables, out, 'testdb', template_path=str(fake))
            self.assertIn('not a readable .xlsx', str(cm.exception))

    def test_bundled_excel_template_matches_the_built_in_default_look(self):
        # excel-template.xlsx is meant to be a working *example* of the
        # built-in styling (regenerated by gen_excel_template.py) — pin it
        # against drift: using it as --excel-template must resolve to the
        # exact same styling as using no template at all.
        try:
            import openpyxl
        except ImportError:
            self.skipTest('openpyxl not installed')
        bundled = ROOT / 'excel-template.xlsx'
        if not bundled.exists():
            self.skipTest('excel-template.xlsx not generated (run gen_excel_template.py)')
        tables = db_tables()
        with tempfile.TemporaryDirectory() as tmp:
            plain = Path(tmp) / 'plain.xlsx'
            templated = Path(tmp) / 'templated.xlsx'
            erd.write_excel(tables, plain, 'testdb')
            erd.write_excel(tables, templated, 'testdb', template_path=str(bundled))
            wp, wt = openpyxl.load_workbook(plain)['Tables'], openpyxl.load_workbook(templated)['Tables']
            for ref in ('A1', 'A3', 'A4', 'A5'):
                cp, ct = wp[ref], wt[ref]
                self.assertEqual(cp.font.b, ct.font.b, ref)
                self.assertEqual(cp.font.sz, ct.font.sz, ref)
                self.assertEqual(
                    cp.font.color.rgb if cp.font.color else None,
                    ct.font.color.rgb if ct.font.color else None, ref)
                self.assertEqual(cp.fill.patternType, ct.fill.patternType, ref)
                if cp.fill.patternType == 'solid':
                    self.assertEqual(cp.fill.fgColor.rgb, ct.fill.fgColor.rgb, ref)
                self.assertEqual(cp.border.left.style, ct.border.left.style, ref)


class TestLoadConfig(unittest.TestCase):
    def _args(self, config=None, no_config=False):
        return SimpleNamespace(config=config, no_config=no_config)

    def test_no_config_no_discovery_returns_empty(self):
        with tempfile.TemporaryDirectory() as tmp:
            orig = os.getcwd()
            os.chdir(tmp)
            self.addCleanup(os.chdir, orig)
            self.assertEqual(erd.load_config(self._args()), {})

    def test_explicit_config_loads_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'my.json'
            path.write_text('{"max_rows": 30, "only": ["a", "b"]}')
            config = erd.load_config(self._args(config=str(path)))
        self.assertEqual(config, {'max_rows': 30, 'only': ['a', 'b']})

    def test_missing_explicit_config_exits(self):
        with self.assertRaises(SystemExit) as cm:
            erd.load_config(self._args(config='/no/such/file.json'))
        self.assertIn('does not exist', str(cm.exception))

    def test_config_and_no_config_together_exits(self):
        with self.assertRaises(SystemExit) as cm:
            erd.load_config(self._args(config='x.json', no_config=True))
        self.assertIn('mutually exclusive', str(cm.exception))

    def test_no_config_skips_auto_discovery(self):
        with tempfile.TemporaryDirectory() as tmp:
            (Path(tmp) / '.erdscope.json').write_text('{"max_rows": 99}')
            orig = os.getcwd()
            os.chdir(tmp)
            self.addCleanup(os.chdir, orig)
            self.assertEqual(erd.load_config(self._args(no_config=True)), {})

    def test_auto_discovers_erdscope_json_in_cwd(self):
        with tempfile.TemporaryDirectory() as tmp:
            (Path(tmp) / '.erdscope.json').write_text('{"max_rows": 42}')
            orig = os.getcwd()
            os.chdir(tmp)
            self.addCleanup(os.chdir, orig)
            self.assertEqual(erd.load_config(self._args()), {'max_rows': 42})

    def test_connection_fields_allowed(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'my.json'
            path.write_text('{"host": "127.0.0.1", "port": 3307, '
                            '"user": "readonly", "database": "myapp_production"}')
            config = erd.load_config(self._args(config=str(path)))
        self.assertEqual(config['host'], '127.0.0.1')
        self.assertEqual(config['database'], 'myapp_production')

    def test_password_key_rejected(self):
        for key in ('password', 'passwd', 'pwd', 'url', 'database_url'):
            with tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / 'my.json'
                path.write_text(json.dumps({key: 'whatever'}))
                with self.assertRaises(SystemExit) as cm:
                    erd.load_config(self._args(config=str(path)))
            self.assertIn(key, str(cm.exception))

    def test_unknown_key_exits(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'my.json'
            path.write_text('{"totally_bogus": 1}')
            with self.assertRaises(SystemExit) as cm:
                erd.load_config(self._args(config=str(path)))
        self.assertIn('totally_bogus', str(cm.exception))

    def test_relations_key_allowed_though_not_in_config_defaults(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'my.json'
            path.write_text('{"relations": []}')
            config = erd.load_config(self._args(config=str(path)))
        self.assertEqual(config, {'relations': []})

    def test_invalid_json_exits_with_clear_message(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'my.json'
            path.write_text('{not valid json')
            with self.assertRaises(SystemExit) as cm:
                erd.load_config(self._args(config=str(path)))
        self.assertIn('JSON', str(cm.exception))

    def test_top_level_must_be_object(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'my.json'
            path.write_text('[1, 2, 3]')
            with self.assertRaises(SystemExit) as cm:
                erd.load_config(self._args(config=str(path)))
        self.assertIn('object', str(cm.exception))

    def test_yaml_config_if_pyyaml_available(self):
        try:
            import yaml  # noqa: F401
        except ImportError:
            self.skipTest('PyYAML not installed')
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'my.yml'
            path.write_text('max_rows: 25\nonly:\n  - a\n  - b\n')
            config = erd.load_config(self._args(config=str(path)))
        self.assertEqual(config, {'max_rows': 25, 'only': ['a', 'b']})


class TestConfigTypeValidation(unittest.TestCase):
    """load_config() checks key names are known; these tests are about the
    separate, later check that each value has the right *type* — a config
    typo here previously reached the browser as a silent failure (a string
    max_rows became a bare JS identifier -> ReferenceError; a string `only`
    got iterated character-by-character by fnmatch, matching everything)."""
    def _load(self, obj):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'my.json'
            path.write_text(json.dumps(obj))
            return erd.load_config(SimpleNamespace(config=str(path), no_config=False))

    def test_max_rows_as_string_rejected(self):
        with self.assertRaises(SystemExit) as cm:
            self._load({'max_rows': 'fifteen'})
        self.assertIn('max_rows', str(cm.exception))

    def test_max_rows_as_bool_rejected(self):
        with self.assertRaises(SystemExit):
            self._load({'max_rows': True})

    def test_max_rows_as_int_accepted(self):
        self.assertEqual(self._load({'max_rows': 30})['max_rows'], 30)

    def test_only_as_bare_string_rejected(self):
        # a bare string instead of a list would otherwise be iterated
        # character-by-character by fnmatch, silently matching every table
        with self.assertRaises(SystemExit) as cm:
            self._load({'only': 'user*'})
        self.assertIn('only', str(cm.exception))

    def test_only_with_non_string_element_rejected(self):
        with self.assertRaises(SystemExit):
            self._load({'only': ['user*', 5]})

    def test_exclude_as_bare_string_rejected(self):
        with self.assertRaises(SystemExit):
            self._load({'exclude': '*_logs'})

    def test_infer_fk_as_string_rejected(self):
        with self.assertRaises(SystemExit) as cm:
            self._load({'infer_fk': 'false'})  # truthy as a Python bool otherwise
        self.assertIn('infer_fk', str(cm.exception))

    def test_infer_fk_as_int_rejected(self):
        with self.assertRaises(SystemExit):
            self._load({'infer_fk': 1})

    def test_infer_fk_bool_accepted(self):
        self.assertEqual(self._load({'infer_fk': True})['infer_fk'], True)

    def test_table_map_as_list_rejected(self):
        with self.assertRaises(SystemExit) as cm:
            self._load({'table_map': ['Widget=crm_widgets']})
        self.assertIn('table_map', str(cm.exception))

    def test_table_map_with_non_string_value_rejected(self):
        with self.assertRaises(SystemExit):
            self._load({'table_map': {'Widget': 5}})

    def test_relations_as_dict_rejected(self):
        with self.assertRaises(SystemExit) as cm:
            self._load({'relations': {'table': 'x'}})
        self.assertIn('relations', str(cm.exception))

    def test_relations_with_non_object_entry_rejected(self):
        with self.assertRaises(SystemExit):
            self._load({'relations': ['not-an-object']})

    def test_output_as_non_string_rejected(self):
        with self.assertRaises(SystemExit) as cm:
            self._load({'output': 123})
        self.assertIn('output', str(cm.exception))


class TestMainConfigIntegration(unittest.TestCase):
    """End-to-end: main() wired to a stubbed parse_mysql, exercising config
    discovery + the CLI/config precedence rule (an explicit CLI flag fully
    replaces the same config key, no merging — including list-valued keys)."""
    def setUp(self):
        orig_parse_mysql = erd.parse_mysql
        orig_argv = sys.argv
        orig_cwd = os.getcwd()
        self.addCleanup(lambda: setattr(erd, 'parse_mysql', orig_parse_mysql))
        self.addCleanup(lambda: setattr(sys, 'argv', orig_argv))
        self.addCleanup(os.chdir, orig_cwd)
        erd.parse_mysql = lambda url: db_tables()
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        os.chdir(self.tmp.name)

    def _run(self, *cli_args):
        sys.argv = ['erd.py', 'mysql://x@localhost/testdb', *cli_args]
        erd.main()

    def _load_output(self, filename='erd.html'):
        html = (Path(self.tmp.name) / filename).read_text()
        m = re.search(r'const DATA = (\{.*?\});\s*\n', html)
        return json.loads(m.group(1))

    def test_config_only_relation_and_filter_applied(self):
        Path('.erdscope.json').write_text(json.dumps({
            'only': ['posts', 'comments'],
            'relations': [{'table': 'comments', 'column': 'post_id',
                          'references': 'posts', 'name': 'thread'}],
        }))
        self._run()
        data = self._load_output()
        self.assertEqual(set(data['tables']), {'posts', 'comments'})
        names = {a['name'] for a in data['tables']['comments']['associations']}
        self.assertIn('thread', names)

    def test_relations_entry_that_is_not_an_object_exits_cleanly(self):
        Path('.erdscope.json').write_text(json.dumps({'relations': ['oops']}))
        with self.assertRaises(SystemExit) as cm:
            self._run()
        self.assertIn('relations', str(cm.exception))

    def test_cli_only_fully_replaces_config_only_not_merged(self):
        Path('.erdscope.json').write_text(json.dumps({'only': ['posts', 'comments', 'users']}))
        self._run('--only', 'tags')
        data = self._load_output()
        self.assertEqual(set(data['tables']), {'tags'})

    def test_cli_output_flag_overrides_config(self):
        Path('.erdscope.json').write_text(json.dumps({'output': 'from_config.html'}))
        self._run('-o', 'from_cli.html')
        self.assertTrue((Path(self.tmp.name) / 'from_cli.html').exists())
        self.assertFalse((Path(self.tmp.name) / 'from_config.html').exists())

    def test_config_output_used_when_cli_omits_it(self):
        Path('.erdscope.json').write_text(json.dumps({'output': 'from_config.html'}))
        self._run()
        self.assertTrue((Path(self.tmp.name) / 'from_config.html').exists())

    def test_no_config_flag_ignores_discovered_file(self):
        Path('.erdscope.json').write_text(json.dumps({'only': ['tags']}))
        self._run('--no-config')
        data = self._load_output()
        self.assertGreater(len(data['tables']), 1)  # unfiltered

    def test_table_map_dict_form_in_config(self):
        # the framework overlay only runs with --models; here we just check
        # the config's table_map dict shape survives the merge loop without
        # needing the CLI "Class=table" string-splitting path
        Path('.erdscope.json').write_text(json.dumps({'table_map': {'Widget': 'crm_widgets'}}))
        self._run('--models', str(FIXTURE))
        data = self._load_output()
        self.assertIn('crm_widgets', data['tables'])
        self.assertNotIn('widgets', data['tables'])

    def test_url_assembled_from_config_connection_fields(self):
        seen = []
        erd.parse_mysql = lambda url: (seen.append(url) or db_tables())
        Path('.erdscope.json').write_text(json.dumps({
            'host': '127.0.0.1', 'port': 3307, 'user': 'readonly', 'database': 'myapp_production',
        }))
        sys.argv = ['erd.py']  # no CLI positional at all — config supplies the connection
        erd.main()
        self.assertEqual(seen, ['mysql://readonly@127.0.0.1:3307/myapp_production'])

    def test_cli_url_wins_over_config_connection_fields(self):
        seen = []
        erd.parse_mysql = lambda url: (seen.append(url) or db_tables())
        Path('.erdscope.json').write_text(json.dumps({
            'host': 'from-config', 'database': 'from_config_db',
        }))
        sys.argv = ['erd.py', 'mysql://cli-wins@example/db']
        erd.main()
        self.assertEqual(seen, ['mysql://cli-wins@example/db'])

    def test_missing_connection_info_exits_with_clear_message(self):
        # Step 8/9: DB is no longer required — with no url AND no --models AND
        # no config.tables there's simply no schema source, which is the error
        # now (previously this said "database URL is required").
        sys.argv = ['erd.py']
        with self.assertRaises(SystemExit) as cm:
            erd.main()
        self.assertIn('no schema input', str(cm.exception))

    def test_yaml_config_carries_connection_fields(self):
        try:
            import yaml  # noqa: F401
        except ImportError:
            self.skipTest('PyYAML not installed')
        seen = []
        erd.parse_mysql = lambda url: (seen.append(url) or db_tables())
        Path('.erdscope.yml').write_text('host: dbhost\nuser: readonly\ndatabase: shop\n')
        sys.argv = ['erd.py']
        erd.main()
        self.assertEqual(seen, ['mysql://readonly@dbhost:3306/shop'])


class TestAssembleConfigUrl(unittest.TestCase):
    """assemble_config_url() builds the mysql:// URL from config connection
    fields — each part is validated against a safe charset first, since
    pasting an unvalidated host/user into the URL string is unsafe: a `/`,
    `@`, or `:` in one of them silently shifts what urlparse reads as the
    host/port/path once the assembled string is re-parsed downstream."""
    def test_no_database_key_returns_none(self):
        self.assertIsNone(erd.assemble_config_url({}))

    def test_full_assembly(self):
        url = erd.assemble_config_url(
            {'host': '10.0.0.5', 'port': 3307, 'user': 'readonly', 'database': 'shop'})
        self.assertEqual(url, 'mysql://readonly@10.0.0.5:3307/shop')

    def test_host_defaults_to_localhost_when_only_database_given(self):
        url = erd.assemble_config_url({'database': 'shop'})
        self.assertEqual(url, 'mysql://127.0.0.1:3306/shop')

    def test_port_defaults_to_3306(self):
        url = erd.assemble_config_url({'host': 'dbhost', 'database': 'shop'})
        self.assertEqual(url, 'mysql://dbhost:3306/shop')

    def test_no_user_omits_auth_segment(self):
        url = erd.assemble_config_url({'host': 'dbhost', 'database': 'shop'})
        self.assertNotIn('@', url)

    def test_user_with_slash_rejected(self):
        with self.assertRaises(SystemExit) as cm:
            erd.assemble_config_url({'host': 'h', 'database': 'shop', 'user': 'read/only'})
        self.assertIn('user', str(cm.exception))

    def test_user_with_at_sign_rejected(self):
        with self.assertRaises(SystemExit):
            erd.assemble_config_url({'host': 'h', 'database': 'shop', 'user': 'a@b'})

    def test_user_with_space_rejected(self):
        with self.assertRaises(SystemExit):
            erd.assemble_config_url({'host': 'h', 'database': 'shop', 'user': 'my user'})

    def test_user_with_password_syntax_rejected_with_specific_message(self):
        with self.assertRaises(SystemExit) as cm:
            erd.assemble_config_url({'host': 'h', 'database': 'shop', 'user': 'readonly:hunter2'})
        self.assertIn('password', str(cm.exception).lower())

    def test_host_with_at_sign_rejected(self):
        # a raw "user@evil"-style host would otherwise silently redistribute
        # across the assembled URL's user/host/port fields when re-parsed
        with self.assertRaises(SystemExit) as cm:
            erd.assemble_config_url({'host': 'x@evil', 'database': 'shop'})
        self.assertIn('host', str(cm.exception))

    def test_host_with_slash_rejected(self):
        with self.assertRaises(SystemExit):
            erd.assemble_config_url({'host': 'x/evil', 'database': 'shop'})

    def test_non_integer_port_rejected(self):
        with self.assertRaises(SystemExit) as cm:
            erd.assemble_config_url({'host': 'h', 'database': 'shop', 'port': 'notaport'})
        self.assertIn('port', str(cm.exception))

    def test_out_of_range_port_rejected(self):
        with self.assertRaises(SystemExit):
            erd.assemble_config_url({'host': 'h', 'database': 'shop', 'port': 99999})

    def test_invalid_database_name_rejected(self):
        with self.assertRaises(SystemExit) as cm:
            erd.assemble_config_url({'host': 'h', 'database': 'shop; drop table x'})
        self.assertIn('database', str(cm.exception))

    def test_blank_database_treated_as_absent_not_as_the_string_none(self):
        # a bare `database:` line in YAML parses to None — must not silently
        # become the literal string "None" in the assembled URL
        self.assertIsNone(erd.assemble_config_url({'database': None}))

    def test_blank_host_falls_back_to_default_not_the_string_none(self):
        url = erd.assemble_config_url({'host': None, 'database': 'shop'})
        self.assertEqual(url, 'mysql://127.0.0.1:3306/shop')

    def test_boolean_port_rejected(self):
        with self.assertRaises(SystemExit):
            erd.assemble_config_url({'host': 'h', 'database': 'shop', 'port': True})

    def test_non_integral_float_port_rejected(self):
        with self.assertRaises(SystemExit):
            erd.assemble_config_url({'host': 'h', 'database': 'shop', 'port': 3306.9})


class TestPostgresQueryRows(unittest.TestCase):
    """postgres_query_rows prefers psycopg (v3), then psycopg2; the last
    resort is the psql CLI wrapping the query in COPY (...) TO STDOUT.
    Mirrors TestPyMysqlErrorHandling: driver failures must exit cleanly,
    and the CLI path must unescape COPY's text format."""
    def setUp(self):
        class FakePgError(Exception):
            pass
        self.FakePgError = FakePgError
        self.fake_psycopg = types.SimpleNamespace(Error=FakePgError, connect=None)
        self.addCleanup(sys.modules.pop, 'psycopg', None)
        sys.modules['psycopg'] = self.fake_psycopg

    def test_connect_failure_exits_cleanly(self):
        def fail_connect(**kw):
            raise self.FakePgError('password authentication failed for user "x"')
        self.fake_psycopg.connect = fail_connect
        with self.assertRaises(SystemExit) as cm:
            erd.postgres_query_rows('postgres://x@127.0.0.1:5432/db', 'SELECT 1')
        self.assertIn('authentication failed', str(cm.exception))

    def test_query_failure_exits_cleanly(self):
        class FakeCursor:
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def execute(self, sql):
                raise self.outer.FakePgError('relation does not exist')
        FakeCursor.outer = self
        class FakeConn:
            def cursor(self): return FakeCursor()
            def close(self): pass
        self.fake_psycopg.connect = lambda **kw: FakeConn()
        with self.assertRaises(SystemExit) as cm:
            erd.postgres_query_rows('postgres://x@127.0.0.1:5432/db', 'SELECT 1')
        self.assertIn('relation does not exist', str(cm.exception))

    def test_none_values_become_empty_strings(self):
        class FakeCursor:
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def execute(self, sql): pass
            def fetchall(self): return [('users', None), (None, 42)]
        class FakeConn:
            def cursor(self): return FakeCursor()
            def close(self): pass
        self.fake_psycopg.connect = lambda **kw: FakeConn()
        rows = erd.postgres_query_rows('postgres://x@h:5432/db', 'SELECT 1')
        self.assertEqual(rows, [('users', ''), ('', '42')])


class TestPostgresCliFallback(unittest.TestCase):
    """The psql path, forced by making both driver imports fail
    (sys.modules[name] = None makes `import name` raise ImportError)."""
    def setUp(self):
        for mod in ('psycopg', 'psycopg2'):
            self.addCleanup(sys.modules.pop, mod, None)
            sys.modules[mod] = None

    def test_missing_psql_exits_with_install_hint(self):
        def no_psql(cmd, **kw):
            raise FileNotFoundError('psql')
        orig = erd.subprocess.run
        erd.subprocess.run = no_psql
        try:
            with self.assertRaises(SystemExit) as cm:
                erd.postgres_query_rows('postgres://x@h:5432/db', 'SELECT 1')
        finally:
            erd.subprocess.run = orig
        self.assertIn('psycopg', str(cm.exception))
        self.assertIn('psql', str(cm.exception))

    def test_copy_output_is_unescaped_and_wrapped_in_copy(self):
        captured = {}
        def fake_run(cmd, **kw):
            captured['cmd'] = cmd
            return types.SimpleNamespace(
                returncode=0,
                stdout='users\tA comment with a\\ttab\n\\N\t\\N\n',
                stderr='')
        orig = erd.subprocess.run
        erd.subprocess.run = fake_run
        try:
            rows = erd.postgres_query_rows('postgres://ro@h:5433/db', 'SELECT x')
        finally:
            erd.subprocess.run = orig
        self.assertEqual(rows, [('users', 'A comment with a\ttab'), ('', '')])
        joined = ' '.join(captured['cmd'])
        self.assertIn('COPY (SELECT x) TO STDOUT', joined)
        self.assertIn('-p 5433', joined)
        self.assertIn('-U ro', joined)


class TestUnescapeCopyField(unittest.TestCase):
    """COPY TO STDOUT text format shares mysql --batch's escape contract
    and adds \\f and \\v."""
    def test_null_marker(self):
        self.assertEqual(erd._unescape_copy_field('\\N'), '')

    def test_formfeed_and_vertical_tab(self):
        self.assertEqual(erd._unescape_copy_field('a\\fb\\vc'), 'a\fb\vc')

    def test_shared_escapes_still_decode(self):
        self.assertEqual(erd._unescape_copy_field('a\\tb\\nc'), 'a\tb\nc')

    def test_escaped_backslash_then_escape_char(self):
        self.assertEqual(erd._unescape_copy_field('a\\\\tb'), 'a\\tb')


class TestPostgresUrlAndSchema(unittest.TestCase):
    def test_missing_database_name_exits(self):
        with self.assertRaises(SystemExit) as cm:
            erd.parse_postgres('postgres://ro@h:5432/')
        self.assertIn('database name', str(cm.exception))

    def test_invalid_schema_param_exits(self):
        with self.assertRaises(SystemExit) as cm:
            erd.parse_postgres('postgres://ro@h:5432/db?schema=bad-name')
        self.assertIn('schema', str(cm.exception))


class TestPostgresConfigUrl(unittest.TestCase):
    """`engine: postgres` in the config switches the assembled URL's scheme
    and default port; mysql stays the default for existing configs."""
    def test_engine_postgres_assembles_postgres_url_with_5432_default(self):
        url = erd.assemble_config_url(
            {'engine': 'postgres', 'user': 'ro', 'host': 'h', 'database': 'shop'})
        self.assertEqual(url, 'postgres://ro@h:5432/shop')

    def test_engine_postgresql_alias_accepted(self):
        url = erd.assemble_config_url({'engine': 'postgresql', 'database': 'shop'})
        self.assertEqual(url, 'postgres://127.0.0.1:5432/shop')

    def test_engine_defaults_to_mysql(self):
        url = erd.assemble_config_url({'database': 'shop'})
        self.assertEqual(url, 'mysql://127.0.0.1:3306/shop')

    def test_unknown_engine_exits(self):
        with self.assertRaises(SystemExit) as cm:
            erd.assemble_config_url({'engine': 'oracle', 'database': 'shop'})
        self.assertIn('engine', str(cm.exception))

    def test_explicit_port_wins_over_engine_default(self):
        url = erd.assemble_config_url(
            {'engine': 'postgres', 'database': 'shop', 'port': 6432})
        self.assertEqual(url, 'postgres://127.0.0.1:6432/shop')


class TestSQLiteAdapter(unittest.TestCase):
    """The SQLite adapter reads a real file via the sqlite3 stdlib module (so
    this is a genuine end-to-end DB test, no container needed) and shapes it
    into the same IR as the other engines."""
    def setUp(self):
        import sqlite3
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.path = os.path.join(self.tmp.name, 'shop.db')
        conn = sqlite3.connect(self.path)
        conn.executescript("""
            CREATE TABLE users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email VARCHAR(255) NOT NULL UNIQUE,
                name TEXT NOT NULL);
            CREATE TABLE posts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                title VARCHAR(200) NOT NULL DEFAULT 'untitled');
            CREATE INDEX idx_posts_user ON posts (user_id);
            CREATE TABLE profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL UNIQUE REFERENCES users(id),
                bio TEXT);
            CREATE VIEW active_users AS SELECT * FROM users;
        """)
        conn.commit()
        conn.close()

    def _url(self):
        return f'sqlite:///{self.path}'

    def test_url_path_parsing(self):
        self.assertEqual(erd.sqlite_path_from_url('sqlite:///rel.db'), 'rel.db')
        self.assertEqual(erd.sqlite_path_from_url('sqlite:////abs/x.db'), '/abs/x.db')

    def test_tables_and_views(self):
        t = erd.parse_sqlite(self._url())
        # the VIEW and internal sqlite_* tables are excluded
        self.assertEqual(sorted(t), ['posts', 'profiles', 'users'])

    def test_columns_types_pk_default_and_autoincrement(self):
        t = erd.parse_sqlite(self._url())
        idc = next(c for c in t['users']['columns'] if c['name'] == 'id')
        self.assertTrue(idc['primary'])
        self.assertFalse(idc['nullable'])         # PK forced NOT NULL
        self.assertEqual(idc['extra'], 'autoincrement')
        self.assertEqual(t['users']['primary_key'], 'id')
        title = next(c for c in t['posts']['columns'] if c['name'] == 'title')
        self.assertEqual(title['type'], 'string')  # VARCHAR -> string shorthand
        self.assertEqual(title['default'], 'untitled')  # quotes stripped

    def test_foreign_keys_and_one_to_one(self):
        t = erd.parse_sqlite(self._url())
        posts = [(a['type'], a['target'], a.get('foreign_key'))
                 for a in t['posts']['associations']]
        self.assertIn(('belongs_to', 'users', 'user_id'), posts)
        # profiles.user_id is UNIQUE -> 1:1, promoted to has_one
        prof = [(a['type'], a['target']) for a in t['profiles']['associations']]
        self.assertEqual(prof, [('has_one', 'users')])

    def test_missing_file_exits(self):
        with self.assertRaises(SystemExit) as cm:
            erd.parse_sqlite('sqlite:///' + os.path.join(self.tmp.name, 'nope.db'))
        self.assertIn('not found', str(cm.exception))

    def test_db_provider_dispatches_to_sqlite(self):
        r = erd.db_provider(self._url())
        self.assertEqual(r['source']['provider'], 'sqlite')
        self.assertIn('users', r['tables'])

    def test_end_to_end_main_generates_html(self):
        out = os.path.join(self.tmp.name, 'out.html')
        argv = sys.argv
        self.addCleanup(lambda: setattr(sys, 'argv', argv))
        sys.argv = ['erd.py', self._url(), '-o', out, '--no-config']
        erd.main()
        with open(out, encoding='utf-8') as f:
            html = f.read()
        self.assertIn('"posts"', html)
        self.assertIn('<title>shop — ERD</title>', html)  # title from the db filename


if __name__ == '__main__':
    unittest.main()
